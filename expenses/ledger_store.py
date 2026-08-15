import os
import tempfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

# Genau zwei Personen -- die Saldo-Berechnung (get_balance) geht fest davon aus, dass jede
# Ausgabe/Rückzahlung zwischen exakt diesen beiden aufgeteilt wird.
PERSONEN = ["Johannes", "Anna"]

TYP_AUSGABE = "Ausgabe"
TYP_GELIEHEN = "Geliehen"
TYP_RUECKZAHLUNG = "Rückzahlung"

_HEADER = ["Datum", "Typ", "Von", "An", "Betrag (EUR)", "Beschreibung"]
_SHEET_NAME = "Ledger"
_OVERVIEW_SHEET_NAME = "Übersicht"
_EUR_FORMAT = '#,##0.00 "€"'


def _other_person(person: str) -> str:
    others = [p for p in PERSONEN if p != person]
    return others[0]


def _atomic_save(workbook: Workbook, path: str) -> None:
    """Speichert nie direkt in-place -- erst in eine temporaere Datei im selben Verzeichnis,
    dann per os.replace() atomar umbenannt. Verhindert eine korrupte/leere Excel-Datei, falls
    der Prozess (Stromausfall, Absturz, systemd-Neustart) genau waehrend eines Schreibvorgangs
    endet -- openpyxl schreibt sonst direkt in die Zieldatei."""
    directory = os.path.dirname(os.path.abspath(path)) or "."
    fd, tmp_path = tempfile.mkstemp(dir=directory, suffix=".xlsx.tmp")
    os.close(fd)
    try:
        workbook.save(tmp_path)
        os.replace(tmp_path, path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


@dataclass
class Entry:
    datum: str
    typ: str
    von: str
    an: str | None
    betrag: float
    beschreibung: str


def init_ledger(path: str) -> None:
    """Legt die Excel-Datei mit Ledger- + Übersichtsblatt an, falls sie noch nicht existiert."""
    file_path = Path(path)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    if file_path.exists():
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET_NAME
    sheet.append(_HEADER)
    _rebuild_overview_sheet(workbook, [])
    _atomic_save(workbook, str(file_path))


def _open_sheet(path: str) -> tuple[Workbook, Worksheet]:
    workbook = load_workbook(path)
    if _SHEET_NAME not in workbook.sheetnames:
        sheet = workbook.active
        sheet.title = _SHEET_NAME
        if sheet.max_row < 1 or sheet.cell(row=1, column=1).value != _HEADER[0]:
            sheet.append(_HEADER)
    return workbook, workbook[_SHEET_NAME]


def _read_entries_indexed(sheet: Worksheet) -> list[tuple[int, Entry]]:
    """Wie _read_entries, aber zusaetzlich mit der Excel-Zeilennummer -- als stabile Referenz
    zum gezielten Bearbeiten/Ueberschreiben einer bestimmten Zeile."""
    entries = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            continue
        datum, typ, von, an, betrag, beschreibung = row[:6]
        entries.append((row_number, Entry(
            datum=str(datum), typ=typ, von=von, an=an, betrag=float(betrag),
            beschreibung=beschreibung or "",
        )))
    return entries


def _read_entries(sheet: Worksheet) -> list[Entry]:
    return [entry for _, entry in _read_entries_indexed(sheet)]


def add_entry(
    path: str, *, typ: str, von: str, betrag: float, beschreibung: str = "", an: str | None = None,
) -> None:
    init_ledger(path)
    workbook, sheet = _open_sheet(path)
    datum = datetime.now(timezone.utc).isoformat(timespec="seconds")
    sheet.append([datum, typ, von, an, betrag, beschreibung])
    _rebuild_overview_sheet(workbook, _read_entries(sheet))
    _atomic_save(workbook, path)


def get_all_entries(path: str) -> list[Entry]:
    init_ledger(path)
    _, sheet = _open_sheet(path)
    return _read_entries(sheet)


def get_all_entries_indexed(path: str) -> list[tuple[int, Entry]]:
    init_ledger(path)
    _, sheet = _open_sheet(path)
    return _read_entries_indexed(sheet)


def get_entry_at(path: str, row_number: int) -> Entry | None:
    init_ledger(path)
    _, sheet = _open_sheet(path)
    for number, entry in _read_entries_indexed(sheet):
        if number == row_number:
            return entry
    return None


def update_entry(
    path: str, row_number: int, *, typ: str, von: str, betrag: float, beschreibung: str = "",
    an: str | None = None,
) -> bool:
    """Ueberschreibt eine bestehende Zeile (Datum der Ersteintragung bleibt erhalten). Gibt
    False zurueck, falls die Zeile zwischenzeitlich verschwunden ist (z.B. per /undo)."""
    init_ledger(path)
    workbook, sheet = _open_sheet(path)
    if sheet.cell(row=row_number, column=1).value is None:
        return False
    sheet.cell(row=row_number, column=2, value=typ)
    sheet.cell(row=row_number, column=3, value=von)
    sheet.cell(row=row_number, column=4, value=an)
    sheet.cell(row=row_number, column=5, value=betrag)
    sheet.cell(row=row_number, column=6, value=beschreibung)
    _rebuild_overview_sheet(workbook, _read_entries(sheet))
    _atomic_save(workbook, path)
    return True


def delete_entry_at(path: str, row_number: int) -> Entry | None:
    """Löscht eine beliebige Zeile. Gibt den gelöschten Eintrag zurück, oder None, falls die
    Zeile nicht (mehr) existiert."""
    init_ledger(path)
    workbook, sheet = _open_sheet(path)
    if sheet.cell(row=row_number, column=1).value is None:
        return None
    datum, typ, von, an, betrag, beschreibung = [
        sheet.cell(row=row_number, column=col).value for col in range(1, 7)
    ]
    sheet.delete_rows(row_number)
    _rebuild_overview_sheet(workbook, _read_entries(sheet))
    _atomic_save(workbook, path)
    return Entry(
        datum=str(datum), typ=typ, von=von, an=an, betrag=float(betrag), beschreibung=beschreibung or "",
    )


def remove_last_entry(path: str) -> Entry | None:
    """Löscht den zuletzt eingetragenen Eintrag (Undo)."""
    init_ledger(path)
    _, sheet = _open_sheet(path)
    if sheet.max_row < 2:
        return None
    return delete_entry_at(path, sheet.max_row)


def get_balance(entries: list[Entry]) -> dict[str, float]:
    """Netto-Saldo pro Person: positiv = wird von der jeweils anderen Person diesen Betrag
    geschuldet. Bei einer Ausgabe (TYP_AUSGABE) wird der Betrag automatisch 50/50 zwischen den
    beiden Personen aufgeteilt (der Zahler hat die Haelfte des Partners vorgestreckt) -- eine
    gemeinsame Ausgabe ist bewusst KEINE Schuld in voller Hoehe. Bei Geliehen (TYP_GELIEHEN,
    "von" leiht "an" Geld, muss zu 100% zurückgezahlt werden) und einer Rückzahlung
    (TYP_RUECKZAHLUNG, "von" zahlt "an" Geld zurück) wird dagegen der VOLLE Betrag verrechnet,
    ohne Split -- beides sind einseitige Geldtransfers, kein gemeinsam getragener Posten.
    Mathematisch ist die Wirkung auf den Saldo fuer beide identisch (balance[von] += Betrag,
    balance[an] -= Betrag); der Typ dient hier nur der lesbaren Unterscheidung im Ledger
    ("geliehen" vs. "zurückgezahlt"). Die Summe beider Salden ist in jedem Fall immer 0."""
    balance = {p: 0.0 for p in PERSONEN}
    for e in entries:
        if e.typ == TYP_AUSGABE:
            other = _other_person(e.von)
            half = e.betrag / 2
            balance[e.von] += half
            balance[other] -= half
        else:
            an = e.an or _other_person(e.von)
            balance[e.von] += e.betrag
            balance[an] -= e.betrag
    return balance


def _format_amount(value: float) -> str:
    return f"{value:.2f}".replace(".", ",") + " €"


def format_saldo(balance: dict[str, float]) -> str:
    a, b = PERSONEN
    amt = round(balance.get(a, 0.0), 2)
    if abs(amt) < 0.01:
        return "Ausgeglichen -- niemand schuldet niemandem etwas."
    if amt > 0:
        return f"{b} schuldet {a}: {_format_amount(amt)}"
    return f"{a} schuldet {b}: {_format_amount(-amt)}"


def _compute_stats(entries: list[Entry]) -> dict:
    """Zentrale Statistik-Berechnung -- von get_summary() (Telegram /stats) UND vom
    Übersichtsblatt (_rebuild_overview_sheet) genutzt, damit beide immer dieselben Zahlen
    zeigen und die Logik nur an einer Stelle gepflegt werden muss."""
    ausgaben = [e for e in entries if e.typ == TYP_AUSGABE]
    geliehen = [e for e in entries if e.typ == TYP_GELIEHEN]
    rueckzahlungen = [e for e in entries if e.typ == TYP_RUECKZAHLUNG]

    by_person_ausgaben_summe = {p: 0.0 for p in PERSONEN}
    by_person_ausgaben_anzahl = {p: 0 for p in PERSONEN}
    for e in ausgaben:
        by_person_ausgaben_summe[e.von] += e.betrag
        by_person_ausgaben_anzahl[e.von] += 1

    by_person_geliehen_summe = {p: 0.0 for p in PERSONEN}
    by_person_geliehen_anzahl = {p: 0 for p in PERSONEN}
    for e in geliehen:
        by_person_geliehen_summe[e.von] += e.betrag
        by_person_geliehen_anzahl[e.von] += 1

    by_person_rueckzahlungen_summe = {p: 0.0 for p in PERSONEN}
    by_person_rueckzahlungen_anzahl = {p: 0 for p in PERSONEN}
    for e in rueckzahlungen:
        by_person_rueckzahlungen_summe[e.von] += e.betrag
        by_person_rueckzahlungen_anzahl[e.von] += 1

    total_ausgaben = sum(e.betrag for e in ausgaben)
    total_geliehen = sum(e.betrag for e in geliehen)
    total_rueckzahlungen = sum(e.betrag for e in rueckzahlungen)
    groesste_ausgabe = max(ausgaben, key=lambda e: e.betrag, default=None)

    return {
        "count": len(entries),
        "count_ausgaben": len(ausgaben),
        "count_geliehen": len(geliehen),
        "count_rueckzahlungen": len(rueckzahlungen),
        "total_ausgaben": total_ausgaben,
        "total_geliehen": total_geliehen,
        "total_rueckzahlungen": total_rueckzahlungen,
        "avg_ausgabe": total_ausgaben / len(ausgaben) if ausgaben else 0.0,
        "by_person_ausgaben_summe": by_person_ausgaben_summe,
        "by_person_ausgaben_anzahl": by_person_ausgaben_anzahl,
        "by_person_geliehen_summe": by_person_geliehen_summe,
        "by_person_geliehen_anzahl": by_person_geliehen_anzahl,
        "by_person_rueckzahlungen_summe": by_person_rueckzahlungen_summe,
        "by_person_rueckzahlungen_anzahl": by_person_rueckzahlungen_anzahl,
        "groesste_ausgabe": groesste_ausgabe,
        "balance": get_balance(entries),
    }


def get_summary(path: str) -> dict:
    return _compute_stats(get_all_entries(path))


def _rebuild_overview_sheet(workbook: Workbook, entries: list[Entry]) -> None:
    """Baut das Übersichtsblatt komplett neu aus den aktuellen Daten -- einfacher und
    robuster als ein bestehendes Blatt inkrementell zu aktualisieren."""
    if _OVERVIEW_SHEET_NAME in workbook.sheetnames:
        del workbook[_OVERVIEW_SHEET_NAME]
    sheet = workbook.create_sheet(_OVERVIEW_SHEET_NAME, 0)

    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)

    sheet["A1"] = "💸 Schulden-Tracker Übersicht"
    sheet["A1"].font = title_font
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 10

    if not entries:
        sheet["A3"] = "Noch keine Einträge vorhanden."
        workbook.active = workbook.sheetnames.index(_OVERVIEW_SHEET_NAME)
        return

    stats = _compute_stats(entries)

    sheet["A3"] = "Gesamt Ausgaben (geteilt):"
    sheet["A3"].font = bold
    sheet["B3"] = stats["total_ausgaben"]
    sheet["B3"].number_format = _EUR_FORMAT

    sheet["A4"] = "Gesamt Geliehen (100%):"
    sheet["A4"].font = bold
    sheet["B4"] = stats["total_geliehen"]
    sheet["B4"].number_format = _EUR_FORMAT

    sheet["A5"] = "Gesamt Rückzahlungen:"
    sheet["A5"].font = bold
    sheet["B5"] = stats["total_rueckzahlungen"]
    sheet["B5"].number_format = _EUR_FORMAT

    sheet["A6"] = "Anzahl Einträge:"
    sheet["A6"].font = bold
    sheet["B6"] = stats["count"]

    sheet["A7"] = "Ø Ausgabe (geteilt):"
    sheet["A7"].font = bold
    sheet["B7"] = stats["avg_ausgabe"]
    sheet["B7"].number_format = _EUR_FORMAT

    groesste = stats["groesste_ausgabe"]
    sheet["A8"] = "Größte Ausgabe (geteilt):"
    sheet["A8"].font = bold
    if groesste is not None:
        sheet["B8"] = groesste.betrag
        sheet["B8"].number_format = _EUR_FORMAT
        sheet["C8"] = f"{groesste.von} — „{groesste.beschreibung}“"

    # Aktueller Saldo
    sheet["A10"] = "Aktueller Saldo"
    sheet["A10"].font = bold
    sheet["A11"] = format_saldo(stats["balance"])
    sheet.merge_cells(start_row=11, start_column=1, end_row=11, end_column=3)
    sheet["A11"].alignment = sheet["A11"].alignment.copy(wrap_text=True)

    # Pro Person -- Ausgaben (geteilt), Geliehen (100%) und Rückzahlungen, je Anzahl + Summe
    sheet["A13"] = "Pro Person"
    sheet["A13"].font = bold
    sheet["A14"] = "Person"
    sheet["B14"] = "Ausgaben (Anzahl)"
    sheet["C14"] = "Ausgaben (Summe)"
    sheet["D14"] = "Geliehen (Anzahl)"
    sheet["E14"] = "Geliehen (Summe)"
    sheet["F14"] = "Rückzahlungen (Anzahl)"
    sheet["G14"] = "Rückzahlungen (Summe)"
    for cell_ref in ("A14", "B14", "C14", "D14", "E14", "F14", "G14"):
        sheet[cell_ref].font = bold
    for col in ("D", "E", "F", "G"):
        sheet.column_dimensions[col].width = 18
    row = 15
    person_start_row = row
    for person in PERSONEN:
        sheet.cell(row=row, column=1, value=person)
        sheet.cell(row=row, column=2, value=stats["by_person_ausgaben_anzahl"][person])
        cell_c = sheet.cell(row=row, column=3, value=stats["by_person_ausgaben_summe"][person])
        cell_c.number_format = _EUR_FORMAT
        sheet.cell(row=row, column=4, value=stats["by_person_geliehen_anzahl"][person])
        cell_e = sheet.cell(row=row, column=5, value=stats["by_person_geliehen_summe"][person])
        cell_e.number_format = _EUR_FORMAT
        sheet.cell(row=row, column=6, value=stats["by_person_rueckzahlungen_anzahl"][person])
        cell_g = sheet.cell(row=row, column=7, value=stats["by_person_rueckzahlungen_summe"][person])
        cell_g.number_format = _EUR_FORMAT
        row += 1
    person_end_row = row - 1

    bar = BarChart()
    bar.title = "Ausgaben pro Person"
    bar.type = "col"
    bar.y_axis.title = "Euro"
    bar.legend = None
    bar_data = Reference(sheet, min_col=3, min_row=person_start_row, max_row=person_end_row)
    bar_labels = Reference(sheet, min_col=1, min_row=person_start_row, max_row=person_end_row)
    bar.add_data(bar_data, titles_from_data=False)
    bar.set_categories(bar_labels)
    bar.height = 9
    bar.width = 14
    sheet.add_chart(bar, "A19")

    workbook.active = workbook.sheetnames.index(_OVERVIEW_SHEET_NAME)
